import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

st.set_page_config(page_title="玉川測量設計 現場データ生成アプリ", layout="wide")
st.title("📐 玉川測量設計 現場データ生成アプリ")

uploaded_file = st.file_uploader("▶ filtered_data.xlsx をアップロードしてください", type="xlsx")

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    st.subheader("データ確認（先頭）")
    st.dataframe(df.head())

    if "地番" not in df.columns or "所在" not in df.columns:
        st.error("必要な列（地番・所在）がありません。")
    else:
        地番一覧 = sorted(df["地番"].dropna().unique(), key=lambda x: int(''.join(filter(str.isdigit, str(x)))))
        col1, col2 = st.columns(2)
        with col1:
            申請地リスト = st.multiselect("✅ 申請地を選択", 地番一覧)
        with col2:
            残り = [x for x in 地番一覧 if x not in 申請地リスト]
            対面地リスト = st.multiselect("🪟 対面地を選択", 残り)

        if st.button("📤 入力シートを出力"):
            def classify(地番):
                if 地番 in 申請地リスト:
                    return "申請地"
                elif 地番 in 対面地リスト:
                    return "対面地"
                else:
                    return "隣接地"

            df["分類"] = df["地番"].apply(classify)
            df["sort_key"] = df["地番"].apply(lambda x: int(''.join(filter(str.isdigit, str(x)))) if pd.notnull(x) else 0)
            df["分類_order"] = pd.Categorical(df["分類"], ["申請地", "隣接地", "対面地"], ordered=True)
            df = df.sort_values(by=["分類_order", "sort_key"])

            df_out = pd.DataFrame()
            df_out["B"] = df["分類"]
            df_out["C"] = df["所在"]
            df_out["D"] = df["地番"]
            df_out["E"] = df["所在"].astype(str) + df["地番"].astype(str)
            df_out["F"] = df.get("地目", "")
            df_out["G"] = df.get("地積", "")
            df_out["H"] = df.get("権利部（甲区）住所", "")
            df_out["I"] = df.get("権利部（甲区）氏名", "")
            df_out["J"] = df.get("権利部（甲区）原因", "")

            wb = load_workbook("現場データ_テンプレート.xlsx")
            ws = wb["入力シート"]

            row_start = 7
            for _, row in df_out.iterrows():
                for col_idx, val in enumerate(row, start=2):
                    ws.cell(row=row_start, column=col_idx).value = val
                row_start += 1

            # ファイル名に申請地の所在を反映
            申請地_df = df[df["分類"] == "申請地"]
            所在名 = str(申請地_df.iloc[0]["所在"]).replace(" ", "").replace("/", "_") if not 申請地_df.empty else "未選択"
            filename = f"現場データ({所在名}).xlsx"

            buffer = io.BytesIO()
            wb.save(buffer)
            st.download_button("📥 入力シート出力", data=buffer.getvalue(),
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
